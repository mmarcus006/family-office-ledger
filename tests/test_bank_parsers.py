"""Tests for bank-specific parsers (CITI, UBS, Morgan Stanley)."""

import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest

from family_office_ledger.parsers.bank_parsers import (
    BankParser,
    BankParserFactory,
    CitiParser,
    MorganStanleyParser,
    ParsedTransaction,
    UBSParser,
)


# ===== ParsedTransaction Tests =====


class TestParsedTransaction:
    """Tests for ParsedTransaction dataclass."""

    def test_create_minimal_transaction(self) -> None:
        """Create transaction with required fields only."""
        txn = ParsedTransaction(
            import_id="abc123",
            date=date(2024, 1, 15),
            description="Test transaction",
            amount=Decimal("100.00"),
            account_number="12345",
            account_name="Checking",
        )

        assert txn.import_id == "abc123"
        assert txn.date == date(2024, 1, 15)
        assert txn.description == "Test transaction"
        assert txn.amount == Decimal("100.00")
        assert txn.account_number == "12345"
        assert txn.account_name == "Checking"
        assert txn.other_party is None
        assert txn.symbol is None
        assert txn.cusip is None
        assert txn.quantity is None
        assert txn.price is None
        assert txn.activity_type is None
        assert txn.raw_data == {}

    def test_create_full_investment_transaction(self) -> None:
        """Create transaction with all investment fields."""
        txn = ParsedTransaction(
            import_id="xyz789",
            date=date(2024, 1, 15),
            description="BOUGHT AAPL",
            amount=Decimal("-15000.00"),
            account_number="67890",
            account_name="Brokerage",
            other_party="Apple Inc",
            symbol="AAPL",
            cusip="037833100",
            quantity=Decimal("100"),
            price=Decimal("150.00"),
            activity_type="BOUGHT",
            raw_data={"original": "data"},
        )

        assert txn.symbol == "AAPL"
        assert txn.cusip == "037833100"
        assert txn.quantity == Decimal("100")
        assert txn.price == Decimal("150.00")
        assert txn.activity_type == "BOUGHT"
        assert txn.raw_data == {"original": "data"}


# ===== CitiParser Tests =====


class TestCitiParser:
    """Tests for CITI CSV parser."""

    def test_parse_standard_citi_csv(self, tmp_path: Path) -> None:
        """Parse standard CITI CSV with =T() account format."""
        csv_content = """Date Range,Account Number,Account Description,Account Nickname,Description,Type,Asset Class,Nominal CCY,Reporting CCY,Amount (Reporting CCY),Amount (Nominal CCY),CUSIP,ISIN,Buy Settlement Amount,Buy Settlement Currency,Category,End of Day Balance(Nominal CCY),End of Day Balance(Reporting CCY),Market Value (Nominal CCY),Principal/Income,Quantity,Security Description,Sell Settlement Amount,Sell Settlement Currency,Settlement Account Description,Settlement Cash Account,Symbol,Transaction Ref # or Deal ID,Value Date,Pending
2026-01-28,=T("XXXXXX9251"),"Citigold Interest Checking",,"ACH ELECTRONIC DEBIT Jan28 06:37a 0000 , CON ED OF NY CECONY ","Cash Withdrawal",,"USD","USD","(217.34)","(217.34)",,,,,"Asset",,,,,,,,,,,"",,,"Yes"
2026-01-27,=T("XXXXXX9251"),"Citigold Interest Checking",,"ACH Electronic Debit, AMERICAN EXPR ACH PMT A0442 1","Cash Withdrawal","Deposits","USD","USD","(146,746.77)","(146,746.77)",,,,,"Asset","70,780.75","70,780.75",,"Both",,,,,,,"",=T("xxxxxx9999"),,"No"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        assert len(result) == 2

        # First transaction
        assert result[0].date == date(2026, 1, 28)
        assert result[0].amount == Decimal(
            "-217.34"
        )  # Parentheses converted to negative
        assert result[0].account_number == "XXXXXX9251"  # Extracted from =T("...")
        assert result[0].account_name == "Citigold Interest Checking"
        assert "CON ED OF NY" in result[0].description
        assert result[0].activity_type == "Cash Withdrawal"

        # Second transaction with large amount and comma
        assert result[1].amount == Decimal("-146746.77")

    def test_citi_extract_account_from_t_format(self, tmp_path: Path) -> None:
        """Extract account number from =T("...") format."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("C92XXX762"),"CGMI Brokerage","Test transaction","Cash Deposit","1000.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        assert result[0].account_number == "C92XXX762"

    def test_citi_handles_plain_account_number(self, tmp_path: Path) -> None:
        """Handle account number without =T() wrapper."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,12345678,"Test Account","Test transaction","Cash Deposit","1000.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        assert result[0].account_number == "12345678"

    def test_citi_parses_investment_fields(self, tmp_path: Path) -> None:
        """Parse CUSIP, quantity, and symbol from CITI investment transactions."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY),CUSIP,Quantity,Symbol
2026-01-15,=T("123456"),"Brokerage","Buy shares","BOUGHT","(5000.00)","037833100","100","AAPL"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        assert result[0].cusip == "037833100"
        assert result[0].quantity == Decimal("100")
        assert result[0].symbol == "AAPL"

    def test_citi_can_parse_recognizes_citi_format(self, tmp_path: Path) -> None:
        """can_parse returns True for CITI CSV files."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("123456"),"Account","Test","Type","100.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        assert parser.can_parse(str(csv_file)) is True

    def test_citi_can_parse_rejects_non_citi(self, tmp_path: Path) -> None:
        """can_parse returns False for non-CITI CSV files."""
        csv_content = """Date,Description,Amount
2024-01-15,DEPOSIT,100.00
"""
        csv_file = tmp_path / "bank.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        assert parser.can_parse(str(csv_file)) is False

    def test_citi_can_parse_rejects_excel(self, tmp_path: Path) -> None:
        """can_parse returns False for Excel files."""
        parser = CitiParser()
        assert parser.can_parse("/path/to/file.xlsx") is False

    def test_citi_generates_unique_import_ids(self, tmp_path: Path) -> None:
        """Each parsed CITI transaction has unique import_id."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("123456"),"Account","Test1","Type","100.00"
2026-01-15,=T("123456"),"Account","Test2","Type","100.00"
2026-01-16,=T("123456"),"Account","Test3","Type","100.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        ids = [txn.import_id for txn in result]
        assert len(ids) == len(set(ids))

    def test_citi_file_not_found_raises(self) -> None:
        """Parsing non-existent file raises FileNotFoundError."""
        parser = CitiParser()
        with pytest.raises(FileNotFoundError):
            parser.parse("/nonexistent/citi.csv")

    def test_citi_empty_file_returns_empty_list(self, tmp_path: Path) -> None:
        """Parsing empty CITI file returns empty list."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        assert result == []

    def test_citi_skips_rows_with_invalid_date(self, tmp_path: Path) -> None:
        """Rows with invalid dates are skipped."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
not-a-date,=T("123456"),"Account","Bad Date","Type","100.00"
2026-01-15,=T("123456"),"Account","Good Date","Type","200.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        assert len(result) == 1
        assert result[0].description == "Good Date"

    def test_citi_skips_rows_with_empty_amount(self, tmp_path: Path) -> None:
        """Rows with empty amounts are skipped."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("123456"),"Account","No Amount","Type",""
2026-01-16,=T("123456"),"Account","Has Amount","Type","50.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        assert len(result) == 1
        assert result[0].description == "Has Amount"

    def test_citi_stores_raw_data(self, tmp_path: Path) -> None:
        """Parsed transaction includes raw row data."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("123456"),"Test Account","Test Desc","Test Type","100.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        assert "Date Range" in result[0].raw_data
        assert result[0].raw_data["Date Range"] == "2026-01-15"


# ===== UBSParser Tests =====


class TestUBSParser:
    """Tests for UBS CSV parser."""

    def test_parse_standard_ubs_csv(self, tmp_path: Path) -> None:
        """Parse standard UBS CSV with filter line on row 1."""
        csv_content = """"Filtered by - Date: 01/01/2025-01/28/2026, Money Market: Exclude"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
V8 03825,01/28/2026,DEBIT CARD,VENMO *Eva Apor,,,Cash,,,-400.00,Atlantic Blue
V8 03825,01/27/2026,WITHDRAWAL,"ACH WITHDRAWAL Deel, Inc.",,,Cash,,,-7757.20,Atlantic Blue
"""
        csv_file = tmp_path / "ubs.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        result = parser.parse(str(csv_file))

        assert len(result) == 2

        # First transaction
        assert result[0].date == date(2026, 1, 28)
        assert result[0].amount == Decimal("-400.00")
        assert result[0].account_number == "V8 03825"
        assert result[0].account_name == "Atlantic Blue"
        assert result[0].description == "VENMO *Eva Apor"
        assert result[0].activity_type == "DEBIT CARD"

        # Second transaction with quoted description
        assert result[1].amount == Decimal("-7757.20")
        assert "Deel, Inc." in result[1].description

    def test_ubs_parses_investment_transactions(self, tmp_path: Path) -> None:
        """Parse UBS investment transactions with symbol, CUSIP, quantity, price."""
        csv_content = """"Filtered by - Date: 01/01/2025-01/28/2026, Money Market: Exclude"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
V8 04232,01/21/2026,BOUGHT,NATL BEVERAGE CORP UNSOLICITED,FIZZ,635017106,Investment,2000.000000,33.32550000,-66751.00,Ikigai
"""
        csv_file = tmp_path / "ubs.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        result = parser.parse(str(csv_file))

        assert result[0].symbol == "FIZZ"
        assert result[0].cusip == "635017106"
        assert result[0].quantity == Decimal("2000.000000")
        assert result[0].price == Decimal("33.32550000")
        assert result[0].activity_type == "BOUGHT"

    def test_ubs_can_parse_recognizes_ubs_format(self, tmp_path: Path) -> None:
        """can_parse returns True for UBS CSV files."""
        csv_content = """"Filtered by - Date: 01/01/2025-01/28/2026"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
V8 03825,01/28/2026,DEBIT CARD,Test,,,Cash,,,-100.00,Test Account
"""
        csv_file = tmp_path / "ubs.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        assert parser.can_parse(str(csv_file)) is True

    def test_ubs_can_parse_rejects_non_ubs(self, tmp_path: Path) -> None:
        """can_parse returns False for non-UBS CSV files."""
        csv_content = """Date,Description,Amount
2024-01-15,DEPOSIT,100.00
"""
        csv_file = tmp_path / "bank.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        assert parser.can_parse(str(csv_file)) is False

    def test_ubs_can_parse_rejects_citi(self, tmp_path: Path) -> None:
        """can_parse returns False for CITI CSV files."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("123456"),"Account","Test","Type","100.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        assert parser.can_parse(str(csv_file)) is False

    def test_ubs_generates_unique_import_ids(self, tmp_path: Path) -> None:
        """Each parsed UBS transaction has unique import_id."""
        csv_content = """"Filtered by - Date: 01/01/2025-01/28/2026"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
V8 03825,01/28/2026,TEST,Test1,,,Cash,,,-100.00,Account
V8 03825,01/28/2026,TEST,Test2,,,Cash,,,-100.00,Account
V8 03825,01/29/2026,TEST,Test3,,,Cash,,,-100.00,Account
"""
        csv_file = tmp_path / "ubs.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        result = parser.parse(str(csv_file))

        ids = [txn.import_id for txn in result]
        assert len(ids) == len(set(ids))

    def test_ubs_file_not_found_raises(self) -> None:
        """Parsing non-existent file raises FileNotFoundError."""
        parser = UBSParser()
        with pytest.raises(FileNotFoundError):
            parser.parse("/nonexistent/ubs.csv")

    def test_ubs_empty_data_returns_empty_list(self, tmp_path: Path) -> None:
        """Parsing UBS file with header only returns empty list."""
        csv_content = """"Filtered by - Date: 01/01/2025-01/28/2026"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
"""
        csv_file = tmp_path / "ubs.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        result = parser.parse(str(csv_file))

        assert result == []

    def test_ubs_uses_activity_as_activity_type(self, tmp_path: Path) -> None:
        """Activity column used as activity_type when Type is empty."""
        csv_content = """"Filtered by - Date: 01/01/2025-01/28/2026"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
V8 03825,01/28/2026,CREDIT,Interest earned,,,,,,-100.00,Account
"""
        csv_file = tmp_path / "ubs.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        result = parser.parse(str(csv_file))

        assert result[0].activity_type == "CREDIT"

    def test_ubs_stores_raw_data(self, tmp_path: Path) -> None:
        """Parsed transaction includes raw row data."""
        csv_content = """"Filtered by - Date: 01/01/2025-01/28/2026"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
V8 03825,01/28/2026,TEST,Test Desc,,,Cash,,,-100.00,Test Account
"""
        csv_file = tmp_path / "ubs.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        result = parser.parse(str(csv_file))

        assert "Account Number" in result[0].raw_data
        assert result[0].raw_data["Account Number"] == "V8 03825"


# ===== MorganStanleyParser Tests =====


class TestMorganStanleyParser:
    """Tests for Morgan Stanley Excel parser."""

    def test_parse_standard_ms_excel(self, tmp_path: Path) -> None:
        """Parse standard Morgan Stanley Excel with header on row 7."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Rows 1-6: Metadata (skipped)
        ws["A1"] = None
        ws["A3"] = "All Activity Types"
        ws["A5"] = "Account Activity for Test from 01/01/2026 - 01/28/2026"

        # Row 7: Header
        headers = [
            "Activity Date",
            "Transaction Date",
            "Account",
            "Activity",
            "Check Number",
            "Card Number",
            "Description",
            "Symbol",
            "Cusip",
            "Category",
            "Memo",
            "Tags",
            "Quantity",
            "Price($)",
            "Amount($)",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        # Row 8: Data
        data = [
            "01/28/2026",
            "01/28/2026",
            "Kaneda - 8231 - 8231",
            "CASH TRANSFER",
            "",
            "",
            "CASH ADJUSTMENT",
            "",
            "",
            "",
            None,
            "",
            0,
            0,
            -110113.71,
        ]
        for col, value in enumerate(data, 1):
            ws.cell(row=8, column=col, value=value)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = MorganStanleyParser()
        result = parser.parse(str(xlsx_file))

        assert len(result) == 1
        assert result[0].date == date(2026, 1, 28)
        assert result[0].amount == Decimal("-110113.71")
        assert (
            result[0].account_number == "8231"
        )  # Extracted from "Kaneda - 8231 - 8231"
        assert result[0].account_name == "Kaneda"
        assert result[0].description == "CASH ADJUSTMENT"
        assert result[0].activity_type == "CASH TRANSFER"

    def test_ms_parses_investment_transactions(self, tmp_path: Path) -> None:
        """Parse Morgan Stanley investment transactions."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Row 7: Header
        headers = [
            "Activity Date",
            "Transaction Date",
            "Account",
            "Activity",
            "Check Number",
            "Card Number",
            "Description",
            "Symbol",
            "Cusip",
            "Category",
            "Memo",
            "Tags",
            "Quantity",
            "Price($)",
            "Amount($)",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        # Row 8: Investment data
        data = [
            "01/27/2026",
            "01/27/2026",
            "Test Account - 6751 - 6751",
            "Interest Income",
            "",
            "",
            "MORGAN STANLEY PRIVATE BANK NA",
            "MSPBNA",
            "061871976",
            "Investment Income",
            None,
            "",
            0,
            0,
            3.24,
        ]
        for col, value in enumerate(data, 1):
            ws.cell(row=8, column=col, value=value)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = MorganStanleyParser()
        result = parser.parse(str(xlsx_file))

        assert result[0].symbol == "MSPBNA"
        assert result[0].cusip == "061871976"
        assert result[0].activity_type == "Interest Income"

    def test_ms_handles_datetime_objects(self, tmp_path: Path) -> None:
        """Handle dates as datetime objects (Excel's default)."""
        from datetime import datetime

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Row 7: Header
        headers = ["Activity Date", "Account", "Activity", "Description", "Amount($)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        # Row 8: Data with datetime object
        ws.cell(row=8, column=1, value=datetime(2026, 1, 28))
        ws.cell(row=8, column=2, value="Account - 1234")
        ws.cell(row=8, column=3, value="TEST")
        ws.cell(row=8, column=4, value="Test")
        ws.cell(row=8, column=5, value=100.0)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = MorganStanleyParser()
        result = parser.parse(str(xlsx_file))

        assert result[0].date == date(2026, 1, 28)

    def test_ms_can_parse_recognizes_ms_format(self, tmp_path: Path) -> None:
        """can_parse returns True for Morgan Stanley Excel files."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Row 7: Header with required columns
        headers = ["Activity Date", "Account", "Activity", "Description", "Amount($)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = MorganStanleyParser()
        assert parser.can_parse(str(xlsx_file)) is True

    def test_ms_can_parse_rejects_csv(self, tmp_path: Path) -> None:
        """can_parse returns False for CSV files."""
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("Date,Amount\n2024-01-15,100")

        parser = MorganStanleyParser()
        assert parser.can_parse(str(csv_file)) is False

    def test_ms_generates_unique_import_ids(self, tmp_path: Path) -> None:
        """Each parsed MS transaction has unique import_id."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Row 7: Header
        headers = ["Activity Date", "Account", "Activity", "Description", "Amount($)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        # Rows 8-10: Data
        for i in range(3):
            ws.cell(row=8 + i, column=1, value="01/28/2026")
            ws.cell(row=8 + i, column=2, value="Account - 1234")
            ws.cell(row=8 + i, column=3, value="TEST")
            ws.cell(row=8 + i, column=4, value=f"Test {i + 1}")
            ws.cell(row=8 + i, column=5, value=100.0)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = MorganStanleyParser()
        result = parser.parse(str(xlsx_file))

        ids = [txn.import_id for txn in result]
        assert len(ids) == len(set(ids))

    def test_ms_file_not_found_raises(self) -> None:
        """Parsing non-existent file raises FileNotFoundError."""
        parser = MorganStanleyParser()
        with pytest.raises(FileNotFoundError):
            parser.parse("/nonexistent/ms.xlsx")

    def test_ms_empty_data_returns_empty_list(self, tmp_path: Path) -> None:
        """Parsing MS file with header only returns empty list."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Row 7: Header only
        headers = ["Activity Date", "Account", "Activity", "Description", "Amount($)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = MorganStanleyParser()
        result = parser.parse(str(xlsx_file))

        assert result == []

    def test_ms_cleans_newlines_in_description(self, tmp_path: Path) -> None:
        """Newlines in description are replaced with spaces."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        headers = ["Activity Date", "Account", "Activity", "Description", "Amount($)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        ws.cell(row=8, column=1, value="01/28/2026")
        ws.cell(row=8, column=2, value="Account - 1234")
        ws.cell(row=8, column=3, value="TEST")
        ws.cell(row=8, column=4, value="Line1\nLine2\nLine3")
        ws.cell(row=8, column=5, value=100.0)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = MorganStanleyParser()
        result = parser.parse(str(xlsx_file))

        assert "\n" not in result[0].description
        assert "Line1 Line2 Line3" == result[0].description

    def test_ms_stores_raw_data(self, tmp_path: Path) -> None:
        """Parsed transaction includes raw row data."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        headers = ["Activity Date", "Account", "Activity", "Description", "Amount($)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        ws.cell(row=8, column=1, value="01/28/2026")
        ws.cell(row=8, column=2, value="Test Account - 1234")
        ws.cell(row=8, column=3, value="TEST ACTIVITY")
        ws.cell(row=8, column=4, value="Test Desc")
        ws.cell(row=8, column=5, value=100.0)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = MorganStanleyParser()
        result = parser.parse(str(xlsx_file))

        assert "activity date" in result[0].raw_data
        assert result[0].raw_data["activity date"] == "01/28/2026"


# ===== BankParserFactory Tests =====


class TestBankParserFactory:
    """Tests for BankParserFactory."""

    def test_factory_returns_citi_parser_for_citi_csv(self, tmp_path: Path) -> None:
        """Factory returns CitiParser for CITI CSV files."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("123456"),"Account","Test","Type","100.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = BankParserFactory.get_parser(str(csv_file))

        assert parser is not None
        assert isinstance(parser, CitiParser)

    def test_factory_returns_ubs_parser_for_ubs_csv(self, tmp_path: Path) -> None:
        """Factory returns UBSParser for UBS CSV files."""
        csv_content = """"Filtered by - Date: 01/01/2025-01/28/2026"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
V8 03825,01/28/2026,TEST,Test,,,Cash,,,-100.00,Account
"""
        csv_file = tmp_path / "ubs.csv"
        csv_file.write_text(csv_content)

        parser = BankParserFactory.get_parser(str(csv_file))

        assert parser is not None
        assert isinstance(parser, UBSParser)

    def test_factory_returns_ms_parser_for_excel(self, tmp_path: Path) -> None:
        """Factory returns MorganStanleyParser for Excel files."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        headers = ["Activity Date", "Account", "Activity", "Description", "Amount($)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = BankParserFactory.get_parser(str(xlsx_file))

        assert parser is not None
        assert isinstance(parser, MorganStanleyParser)

    def test_factory_returns_none_for_unknown_format(self, tmp_path: Path) -> None:
        """Factory returns None for unrecognized file formats."""
        csv_content = """Date,Description,Amount
2024-01-15,DEPOSIT,100.00
"""
        csv_file = tmp_path / "generic.csv"
        csv_file.write_text(csv_content)

        parser = BankParserFactory.get_parser(str(csv_file))

        assert parser is None

    def test_factory_parse_citi_csv(self, tmp_path: Path) -> None:
        """Factory.parse works with CITI CSV files."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("123456"),"Account","Test Transaction","Type","100.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        result = BankParserFactory.parse(str(csv_file))

        assert len(result) == 1
        assert result[0].description == "Test Transaction"

    def test_factory_parse_raises_for_unknown_format(self, tmp_path: Path) -> None:
        """Factory.parse raises ValueError for unknown formats."""
        csv_content = """Date,Description,Amount
2024-01-15,DEPOSIT,100.00
"""
        csv_file = tmp_path / "generic.csv"
        csv_file.write_text(csv_content)

        with pytest.raises(ValueError, match="No parser available"):
            BankParserFactory.parse(str(csv_file))

    def test_factory_parse_raises_for_missing_file(self) -> None:
        """Factory.parse raises FileNotFoundError for missing files."""
        with pytest.raises(FileNotFoundError):
            BankParserFactory.parse("/nonexistent/file.csv")


# ===== Integration Tests with Real Files =====


class TestRealFileIntegration:
    """Integration tests with real bank files (if available)."""

    CITI_FILE = "/mnt/c/Users/Miller/Downloads/Transactions_CITI_01_28_2026_YTD.csv"
    UBS_FILE = "/mnt/c/Users/Miller/Downloads/Transactions_UBS_01_28_2026_YTD.csv"
    MS_FILE = "/mnt/c/Users/Miller/Downloads/Transactions_MS_01_28_2026_YTD.xlsx"

    @pytest.mark.skipif(
        not Path(CITI_FILE).exists(),
        reason="CITI test file not available",
    )
    def test_parse_real_citi_file(self) -> None:
        """Parse real CITI CSV file."""
        parser = CitiParser()
        result = parser.parse(self.CITI_FILE)

        # Should have multiple transactions
        assert len(result) > 0

        # Verify first transaction has expected structure
        txn = result[0]
        assert txn.import_id is not None and len(txn.import_id) == 16
        assert isinstance(txn.date, date)
        assert isinstance(txn.amount, Decimal)
        assert txn.account_number != ""  # Account should be extracted
        assert "=T(" not in txn.account_number  # Should not contain =T wrapper

    @pytest.mark.skipif(
        not Path(UBS_FILE).exists(),
        reason="UBS test file not available",
    )
    def test_parse_real_ubs_file(self) -> None:
        """Parse real UBS CSV file."""
        parser = UBSParser()
        result = parser.parse(self.UBS_FILE)

        # Should have multiple transactions
        assert len(result) > 0

        # Verify first transaction has expected structure
        txn = result[0]
        assert txn.import_id is not None and len(txn.import_id) == 16
        assert isinstance(txn.date, date)
        assert isinstance(txn.amount, Decimal)
        assert txn.account_name != ""  # Friendly account name should be present

    @pytest.mark.skipif(
        not Path(MS_FILE).exists(),
        reason="Morgan Stanley test file not available",
    )
    def test_parse_real_ms_file(self) -> None:
        """Parse real Morgan Stanley Excel file."""
        parser = MorganStanleyParser()
        result = parser.parse(self.MS_FILE)

        # Should have multiple transactions
        assert len(result) > 0

        # Verify first transaction has expected structure
        txn = result[0]
        assert txn.import_id is not None and len(txn.import_id) == 16
        assert isinstance(txn.date, date)
        assert isinstance(txn.amount, Decimal)
        assert txn.account_number != ""  # Account number should be extracted

    @pytest.mark.skipif(
        not Path(CITI_FILE).exists(),
        reason="CITI test file not available",
    )
    def test_factory_auto_detects_citi(self) -> None:
        """Factory correctly identifies CITI file."""
        parser = BankParserFactory.get_parser(self.CITI_FILE)
        assert isinstance(parser, CitiParser)

    @pytest.mark.skipif(
        not Path(UBS_FILE).exists(),
        reason="UBS test file not available",
    )
    def test_factory_auto_detects_ubs(self) -> None:
        """Factory correctly identifies UBS file."""
        parser = BankParserFactory.get_parser(self.UBS_FILE)
        assert isinstance(parser, UBSParser)

    @pytest.mark.skipif(
        not Path(MS_FILE).exists(),
        reason="Morgan Stanley test file not available",
    )
    def test_factory_auto_detects_ms(self) -> None:
        """Factory correctly identifies Morgan Stanley file."""
        parser = BankParserFactory.get_parser(self.MS_FILE)
        assert isinstance(parser, MorganStanleyParser)


# ===== Edge Case Tests =====


class TestEdgeCases:
    """Edge case tests for all parsers."""

    def test_citi_handles_empty_cusip_and_symbol(self, tmp_path: Path) -> None:
        """Empty CUSIP and symbol fields don't cause issues."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY),CUSIP,Symbol
2026-01-15,=T("123456"),"Account","Cash transaction","Type","100.00",,
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result = parser.parse(str(csv_file))

        assert result[0].cusip is None
        assert result[0].symbol is None

    def test_ubs_handles_negative_quantity(self, tmp_path: Path) -> None:
        """Negative quantity (short sales) handled correctly."""
        csv_content = """"Filtered by - Date: 01/01/2025-01/28/2026"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
V8 03825,01/28/2026,SOLD SHORT,Test Stock,TEST,123456789,Investment,-100,50.00,5000.00,Account
"""
        csv_file = tmp_path / "ubs.csv"
        csv_file.write_text(csv_content)

        parser = UBSParser()
        result = parser.parse(str(csv_file))

        assert result[0].quantity == Decimal("-100")

    def test_ms_handles_zero_quantity_and_price(self, tmp_path: Path) -> None:
        """Zero quantity and price for cash transactions."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        headers = [
            "Activity Date",
            "Account",
            "Activity",
            "Description",
            "Quantity",
            "Price($)",
            "Amount($)",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)

        ws.cell(row=8, column=1, value="01/28/2026")
        ws.cell(row=8, column=2, value="Account - 1234")
        ws.cell(row=8, column=3, value="CASH TRANSFER")
        ws.cell(row=8, column=4, value="Transfer")
        ws.cell(row=8, column=5, value=0)
        ws.cell(row=8, column=6, value=0)
        ws.cell(row=8, column=7, value=-1000.0)

        xlsx_file = tmp_path / "ms.xlsx"
        wb.save(xlsx_file)
        wb.close()

        parser = MorganStanleyParser()
        result = parser.parse(str(xlsx_file))

        assert result[0].quantity == Decimal("0")
        assert result[0].price == Decimal("0")
        assert result[0].amount == Decimal("-1000.0")

    def test_deterministic_import_ids(self, tmp_path: Path) -> None:
        """Import IDs are deterministic across multiple parses."""
        csv_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("123456"),"Account","Test Transaction","Type","100.00"
"""
        csv_file = tmp_path / "citi.csv"
        csv_file.write_text(csv_content)

        parser = CitiParser()
        result1 = parser.parse(str(csv_file))
        result2 = parser.parse(str(csv_file))

        # Same file should produce same import IDs
        assert result1[0].import_id == result2[0].import_id

    def test_all_parsers_return_parsed_transaction_type(self, tmp_path: Path) -> None:
        """All parsers return list of ParsedTransaction objects."""
        # CITI
        citi_content = """Date Range,Account Number,Account Description,Description,Type,Amount (Reporting CCY)
2026-01-15,=T("123456"),"Account","Test","Type","100.00"
"""
        citi_file = tmp_path / "citi.csv"
        citi_file.write_text(citi_content)

        citi_result = CitiParser().parse(str(citi_file))
        assert all(isinstance(txn, ParsedTransaction) for txn in citi_result)

        # UBS
        ubs_content = """"Filtered by - Date: 01/01/2025-01/28/2026"
Account Number,Date,Activity,Description,Symbol,Cusip,Type,Quantity,Price,Amount,Friendly Account Name
V8 03825,01/28/2026,TEST,Test,,,Cash,,,-100.00,Account
"""
        ubs_file = tmp_path / "ubs.csv"
        ubs_file.write_text(ubs_content)

        ubs_result = UBSParser().parse(str(ubs_file))
        assert all(isinstance(txn, ParsedTransaction) for txn in ubs_result)

        # Morgan Stanley
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        headers = ["Activity Date", "Account", "Activity", "Description", "Amount($)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=7, column=col, value=header)
        ws.cell(row=8, column=1, value="01/28/2026")
        ws.cell(row=8, column=2, value="Account - 1234")
        ws.cell(row=8, column=3, value="TEST")
        ws.cell(row=8, column=4, value="Test")
        ws.cell(row=8, column=5, value=100.0)

        ms_file = tmp_path / "ms.xlsx"
        wb.save(ms_file)
        wb.close()

        ms_result = MorganStanleyParser().parse(str(ms_file))
        assert all(isinstance(txn, ParsedTransaction) for txn in ms_result)
