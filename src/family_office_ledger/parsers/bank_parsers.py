"""Bank-specific parsers for CITI, UBS, and Morgan Stanley statement formats."""

import csv
import hashlib
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]


@dataclass
class ParsedTransaction:
    """Standardized transaction from any bank parser.

    Attributes:
        import_id: Unique ID for deduplication (SHA256 hash)
        date: Transaction date
        description: Bank description
        amount: Transaction amount (positive=credit, negative=debit)
        account_number: Bank account number
        account_name: Friendly account name
        other_party: Counterparty name (if available)
        symbol: Security symbol (if investment)
        cusip: CUSIP (if investment)
        quantity: Shares (if investment)
        price: Price per share (if investment)
        activity_type: Bank's activity type (BOUGHT, SOLD, etc.)
        raw_data: Original row for debugging
    """

    import_id: str
    date: date
    description: str
    amount: Decimal
    account_number: str
    account_name: str
    other_party: str | None = None
    symbol: str | None = None
    cusip: str | None = None
    quantity: Decimal | None = None
    price: Decimal | None = None
    activity_type: str | None = None
    raw_data: dict[str, Any] = field(default_factory=dict)


class BankParser(ABC):
    """Abstract base class for bank-specific parsers."""

    # Date formats to try when parsing
    DATE_FORMATS = [
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%m-%d-%Y",
        "%d-%m-%Y",
    ]

    @abstractmethod
    def parse(self, file_path: str) -> list[ParsedTransaction]:
        """Parse a file and return standardized transactions.

        Args:
            file_path: Path to the file to parse.

        Returns:
            List of ParsedTransaction objects.

        Raises:
            FileNotFoundError: If the file does not exist.
        """
        ...

    @abstractmethod
    def can_parse(self, file_path: str) -> bool:
        """Check if this parser can handle the given file.

        Args:
            file_path: Path to the file to check.

        Returns:
            True if this parser can handle the file.
        """
        ...

    def _parse_date(self, date_str: str, date_format: str | None = None) -> date | None:
        """Parse a date string using various formats.

        Args:
            date_str: Date string to parse.
            date_format: Optional specific date format to use.

        Returns:
            Parsed date or None if parsing fails.
        """
        if not date_str:
            return None

        date_str = date_str.strip()

        if date_format:
            try:
                return datetime.strptime(date_str, date_format).date()
            except ValueError:
                return None

        for fmt in self.DATE_FORMATS:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        return None

    def _parse_decimal(self, value: str | float | int | None) -> Decimal | None:
        """Parse a value to Decimal, handling currency symbols and formatting.

        Args:
            value: Value to parse (string, float, int, or None).

        Returns:
            Decimal or None if parsing fails.
        """
        if value is None:
            return None

        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value))
            except InvalidOperation:
                return None

        # String processing
        cleaned = str(value).strip()
        if not cleaned:
            return None

        # Remove currency symbols and thousands separators
        cleaned = cleaned.replace("$", "").replace(",", "").replace(" ", "")

        # Handle parentheses for negative numbers
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]

        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None

    def _generate_import_id(
        self, file_path: str, row_num: int, txn_date: date, amount: Decimal
    ) -> str:
        """Generate a unique import ID for a transaction.

        Args:
            file_path: Source file path.
            row_num: Row number in the file.
            txn_date: Transaction date.
            amount: Transaction amount.

        Returns:
            Unique import ID string (first 16 chars of SHA256 hash).
        """
        hash_input = f"{file_path}:{row_num}:{txn_date.isoformat()}:{amount}"
        return hashlib.sha256(hash_input.encode()).hexdigest()[:16]


class CitiParser(BankParser):
    """Parser for CITI CSV bank statement exports.

    Handles CITI-specific format including:
    - Account numbers in =T("...") format
    - Amounts with parentheses for negatives
    - Date format: YYYY-MM-DD
    """

    # CITI column mappings (case-insensitive)
    DATE_COLUMN = "date range"
    ACCOUNT_COLUMN = "account number"
    ACCOUNT_NAME_COLUMN = "account description"
    DESCRIPTION_COLUMN = "description"
    AMOUNT_COLUMN = "amount (reporting ccy)"
    TYPE_COLUMN = "type"
    CUSIP_COLUMN = "cusip"
    QUANTITY_COLUMN = "quantity"
    SYMBOL_COLUMN = "symbol"

    # Regex to extract account from =T("XXXXXX9251")
    ACCOUNT_PATTERN = re.compile(r'=T\("([^"]+)"\)')

    def can_parse(self, file_path: str) -> bool:
        """Check if this is a CITI CSV file."""
        path = Path(file_path)
        if path.suffix.lower() != ".csv":
            return False

        try:
            with open(path, newline="", encoding="utf-8-sig") as csvfile:
                reader = csv.DictReader(csvfile)
                if reader.fieldnames is None:
                    return False
                fieldnames_lower = [f.lower().strip() for f in reader.fieldnames]
                # Check for CITI-specific columns
                return (
                    self.DATE_COLUMN in fieldnames_lower
                    and self.ACCOUNT_COLUMN in fieldnames_lower
                    and self.AMOUNT_COLUMN in fieldnames_lower
                )
        except Exception:
            return False

    def parse(self, file_path: str) -> list[ParsedTransaction]:
        """Parse a CITI CSV file."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"CSV file not found: {file_path}")

        transactions: list[ParsedTransaction] = []

        with open(path, newline="", encoding="utf-8-sig") as csvfile:
            reader = csv.DictReader(csvfile)
            if reader.fieldnames is None:
                return []

            # Build column mapping (case-insensitive)
            col_map = {f.lower().strip(): f for f in reader.fieldnames}

            for row_num, row in enumerate(reader, start=1):
                txn = self._parse_row(row, col_map, file_path, row_num)
                if txn is not None:
                    transactions.append(txn)

        return transactions

    def _parse_row(
        self,
        row: dict[str, str],
        col_map: dict[str, str],
        file_path: str,
        row_num: int,
    ) -> ParsedTransaction | None:
        """Parse a single CITI CSV row."""
        # Get column names from mapping
        date_col = col_map.get(self.DATE_COLUMN)
        account_col = col_map.get(self.ACCOUNT_COLUMN)
        account_name_col = col_map.get(self.ACCOUNT_NAME_COLUMN)
        desc_col = col_map.get(self.DESCRIPTION_COLUMN)
        amount_col = col_map.get(self.AMOUNT_COLUMN)
        type_col = col_map.get(self.TYPE_COLUMN)
        cusip_col = col_map.get(self.CUSIP_COLUMN)
        quantity_col = col_map.get(self.QUANTITY_COLUMN)
        symbol_col = col_map.get(self.SYMBOL_COLUMN)

        # Parse date (required)
        date_str = row.get(date_col, "").strip() if date_col else ""
        if not date_str:
            return None

        txn_date = self._parse_date(date_str)
        if txn_date is None:
            return None

        # Parse amount (required)
        amount_str = row.get(amount_col, "").strip() if amount_col else ""
        amount = self._parse_decimal(amount_str)
        if amount is None:
            return None

        # Extract account number from =T("...") format
        raw_account = row.get(account_col, "") if account_col else ""
        account_number = self._extract_account(raw_account)

        # Get other fields
        account_name = row.get(account_name_col, "").strip() if account_name_col else ""
        description = row.get(desc_col, "").strip() if desc_col else ""
        activity_type = row.get(type_col, "").strip() if type_col else None
        cusip = row.get(cusip_col, "").strip() if cusip_col else None
        symbol = row.get(symbol_col, "").strip() if symbol_col else None

        # Parse quantity
        quantity_str = row.get(quantity_col, "") if quantity_col else ""
        quantity = self._parse_decimal(quantity_str)

        # Generate import ID
        import_id = self._generate_import_id(file_path, row_num, txn_date, amount)

        return ParsedTransaction(
            import_id=import_id,
            date=txn_date,
            description=description,
            amount=amount,
            account_number=account_number,
            account_name=account_name,
            other_party=None,
            symbol=symbol if symbol else None,
            cusip=cusip if cusip else None,
            quantity=quantity,
            price=None,  # CITI doesn't have price per share in this format
            activity_type=activity_type if activity_type else None,
            raw_data=dict(row),
        )

    def _extract_account(self, raw_value: str) -> str:
        """Extract account number from =T("...") format.

        Args:
            raw_value: Raw account value like =T("XXXXXX9251")

        Returns:
            Extracted account number or raw value if no match.
        """
        if not raw_value:
            return ""

        match = self.ACCOUNT_PATTERN.search(raw_value)
        if match:
            return match.group(1)
        return raw_value.strip()


class UBSParser(BankParser):
    """Parser for UBS CSV bank statement exports.

    Handles UBS-specific format including:
    - Filter metadata line on row 1 (skip it)
    - Header on row 2
    - Date format: MM/DD/YYYY
    """

    # UBS column mappings (case-insensitive)
    ACCOUNT_COLUMN = "account number"
    DATE_COLUMN = "date"
    ACTIVITY_COLUMN = "activity"
    DESCRIPTION_COLUMN = "description"
    SYMBOL_COLUMN = "symbol"
    CUSIP_COLUMN = "cusip"
    TYPE_COLUMN = "type"
    QUANTITY_COLUMN = "quantity"
    PRICE_COLUMN = "price"
    AMOUNT_COLUMN = "amount"
    ACCOUNT_NAME_COLUMN = "friendly account name"

    def can_parse(self, file_path: str) -> bool:
        """Check if this is a UBS CSV file."""
        path = Path(file_path)
        if path.suffix.lower() != ".csv":
            return False

        try:
            with open(path, newline="", encoding="utf-8-sig") as csvfile:
                # Read first line to check for filter metadata
                first_line = csvfile.readline()
                if not first_line.lower().startswith('"filtered by'):
                    return False

                # Read second line for headers
                reader = csv.DictReader(csvfile)
                if reader.fieldnames is None:
                    return False
                fieldnames_lower = [f.lower().strip() for f in reader.fieldnames]
                return (
                    self.ACCOUNT_COLUMN in fieldnames_lower
                    and self.DATE_COLUMN in fieldnames_lower
                    and self.AMOUNT_COLUMN in fieldnames_lower
                    and self.ACCOUNT_NAME_COLUMN in fieldnames_lower
                )
        except Exception:
            return False

    def parse(self, file_path: str) -> list[ParsedTransaction]:
        """Parse a UBS CSV file."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"CSV file not found: {file_path}")

        transactions: list[ParsedTransaction] = []

        with open(path, newline="", encoding="utf-8-sig") as csvfile:
            # Skip the first line (filter metadata)
            csvfile.readline()

            reader = csv.DictReader(csvfile)
            if reader.fieldnames is None:
                return []

            # Build column mapping (case-insensitive)
            col_map = {f.lower().strip(): f for f in reader.fieldnames}

            for row_num, row in enumerate(
                reader, start=2
            ):  # Start at 2 since we skipped row 1
                txn = self._parse_row(row, col_map, file_path, row_num)
                if txn is not None:
                    transactions.append(txn)

        return transactions

    def _parse_row(
        self,
        row: dict[str, str],
        col_map: dict[str, str],
        file_path: str,
        row_num: int,
    ) -> ParsedTransaction | None:
        """Parse a single UBS CSV row."""
        # Get column names from mapping
        account_col = col_map.get(self.ACCOUNT_COLUMN)
        date_col = col_map.get(self.DATE_COLUMN)
        activity_col = col_map.get(self.ACTIVITY_COLUMN)
        desc_col = col_map.get(self.DESCRIPTION_COLUMN)
        symbol_col = col_map.get(self.SYMBOL_COLUMN)
        cusip_col = col_map.get(self.CUSIP_COLUMN)
        type_col = col_map.get(self.TYPE_COLUMN)
        quantity_col = col_map.get(self.QUANTITY_COLUMN)
        price_col = col_map.get(self.PRICE_COLUMN)
        amount_col = col_map.get(self.AMOUNT_COLUMN)
        account_name_col = col_map.get(self.ACCOUNT_NAME_COLUMN)

        # Parse date (required)
        date_str = row.get(date_col, "").strip() if date_col else ""
        if not date_str:
            return None

        txn_date = self._parse_date(date_str)
        if txn_date is None:
            return None

        # Parse amount (required)
        amount_str = row.get(amount_col, "").strip() if amount_col else ""
        amount = self._parse_decimal(amount_str)
        if amount is None:
            return None

        # Get other fields
        account_number = row.get(account_col, "").strip() if account_col else ""
        account_name = row.get(account_name_col, "").strip() if account_name_col else ""
        description = row.get(desc_col, "").strip() if desc_col else ""
        activity = row.get(activity_col, "").strip() if activity_col else ""
        type_val = row.get(type_col, "").strip() if type_col else ""
        cusip = row.get(cusip_col, "").strip() if cusip_col else None
        symbol = row.get(symbol_col, "").strip() if symbol_col else None

        # Parse quantity and price
        quantity_str = row.get(quantity_col, "") if quantity_col else ""
        price_str = row.get(price_col, "") if price_col else ""
        quantity = self._parse_decimal(quantity_str)
        price = self._parse_decimal(price_str)

        # Prefer Activity column (BOUGHT, SOLD, DEBIT CARD, etc.) over Type column (Cash, Investment)
        activity_type: str | None = (
            activity if activity else (type_val if type_val else None)
        )

        # Generate import ID
        import_id = self._generate_import_id(file_path, row_num, txn_date, amount)

        return ParsedTransaction(
            import_id=import_id,
            date=txn_date,
            description=description,
            amount=amount,
            account_number=account_number,
            account_name=account_name,
            other_party=None,
            symbol=symbol if symbol else None,
            cusip=cusip if cusip else None,
            quantity=quantity,
            price=price,
            activity_type=activity_type if activity_type else None,
            raw_data=dict(row),
        )


class MorganStanleyParser(BankParser):
    """Parser for Morgan Stanley Excel statement exports.

    Handles Morgan Stanley-specific format including:
    - Rows 1-6: Metadata (skip)
    - Row 7: Header
    - Row 8+: Data
    - Date format: MM/DD/YYYY
    - Uses openpyxl for Excel parsing
    """

    # Morgan Stanley column mappings (case-insensitive)
    DATE_COLUMN = "activity date"
    TRANSACTION_DATE_COLUMN = "transaction date"
    ACCOUNT_COLUMN = "account"
    ACTIVITY_COLUMN = "activity"
    DESCRIPTION_COLUMN = "description"
    SYMBOL_COLUMN = "symbol"
    CUSIP_COLUMN = "cusip"
    QUANTITY_COLUMN = "quantity"
    PRICE_COLUMN = "price($)"
    AMOUNT_COLUMN = "amount($)"

    HEADER_ROW = 7
    DATA_START_ROW = 8

    def can_parse(self, file_path: str) -> bool:
        """Check if this is a Morgan Stanley Excel file."""
        path = Path(file_path)
        if path.suffix.lower() not in (".xlsx", ".xls"):
            return False

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                return False

            # Check row 7 for expected headers
            header_row = [
                str(cell.value).lower().strip() if cell.value else ""
                for cell in ws[self.HEADER_ROW]
            ]
            wb.close()

            return (
                self.DATE_COLUMN in header_row
                and self.ACCOUNT_COLUMN in header_row
                and self.AMOUNT_COLUMN in header_row
            )
        except Exception:
            return False

    def parse(self, file_path: str) -> list[ParsedTransaction]:
        """Parse a Morgan Stanley Excel file."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        transactions: list[ParsedTransaction] = []

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return []

        try:
            # Read header row (row 7)
            header_row = list(
                ws.iter_rows(min_row=self.HEADER_ROW, max_row=self.HEADER_ROW)
            )[0]
            headers = [
                str(cell.value).lower().strip() if cell.value else ""
                for cell in header_row
            ]

            # Build column index mapping
            col_indices: dict[str, int] = {}
            for idx, header in enumerate(headers):
                col_indices[header] = idx

            # Read data rows (row 8+)
            for row_num, row in enumerate(
                ws.iter_rows(min_row=self.DATA_START_ROW),
                start=self.DATA_START_ROW,
            ):
                txn = self._parse_row(row, col_indices, file_path, row_num)
                if txn is not None:
                    transactions.append(txn)
        finally:
            wb.close()

        return transactions

    def _parse_row(
        self,
        row: tuple,  # type: ignore[type-arg]
        col_indices: dict[str, int],
        file_path: str,
        row_num: int,
    ) -> ParsedTransaction | None:
        """Parse a single Morgan Stanley Excel row."""

        def get_cell_value(col_name: str) -> Any:
            idx = col_indices.get(col_name)
            if idx is not None and idx < len(row):
                return row[idx].value
            return None

        # Parse date (required) - prefer activity date
        date_val = get_cell_value(self.DATE_COLUMN)
        if date_val is None:
            date_val = get_cell_value(self.TRANSACTION_DATE_COLUMN)

        if date_val is None:
            return None

        # Handle date as string or datetime object
        txn_date: date
        if isinstance(date_val, datetime):
            txn_date = date_val.date()
        elif isinstance(date_val, date):
            txn_date = date_val
        elif isinstance(date_val, str):
            parsed_date = self._parse_date(date_val)
            if parsed_date is None:
                return None
            txn_date = parsed_date
        else:
            return None

        # Parse amount (required)
        amount_val = get_cell_value(self.AMOUNT_COLUMN)
        amount = self._parse_decimal(amount_val)
        if amount is None:
            return None

        # Get account info
        account_raw = get_cell_value(self.ACCOUNT_COLUMN)
        account_str = str(account_raw).strip() if account_raw else ""
        # Extract account number from "Entity Name - XXXX - XXXX" format
        account_number, account_name = self._parse_account(account_str)

        # Get other fields
        description_val = get_cell_value(self.DESCRIPTION_COLUMN)
        description = str(description_val).strip() if description_val else ""
        # Clean up newlines in description
        description = description.replace("\n", " ").strip()

        activity_val = get_cell_value(self.ACTIVITY_COLUMN)
        activity_type = str(activity_val).strip() if activity_val else None

        symbol_val = get_cell_value(self.SYMBOL_COLUMN)
        symbol = str(symbol_val).strip() if symbol_val else None

        cusip_val = get_cell_value(self.CUSIP_COLUMN)
        cusip = str(cusip_val).strip() if cusip_val else None

        # Parse quantity and price
        quantity_val = get_cell_value(self.QUANTITY_COLUMN)
        price_val = get_cell_value(self.PRICE_COLUMN)
        quantity = self._parse_decimal(quantity_val)
        price = self._parse_decimal(price_val)

        # Generate import ID
        import_id = self._generate_import_id(file_path, row_num, txn_date, amount)

        # Build raw_data dict
        raw_data: dict[str, Any] = {}
        for col_name, idx in col_indices.items():
            if idx < len(row) and col_name:
                raw_data[col_name] = row[idx].value

        return ParsedTransaction(
            import_id=import_id,
            date=txn_date,
            description=description,
            amount=amount,
            account_number=account_number,
            account_name=account_name,
            other_party=None,
            symbol=symbol if symbol else None,
            cusip=cusip if cusip else None,
            quantity=quantity,
            price=price,
            activity_type=activity_type if activity_type else None,
            raw_data=raw_data,
        )

    def _parse_account(self, account_str: str) -> tuple[str, str]:
        """Parse account string in format 'Entity Name - XXXX - XXXX'.

        Args:
            account_str: Full account string.

        Returns:
            Tuple of (account_number, account_name).
        """
        if not account_str:
            return "", ""

        parts = account_str.split(" - ")
        if len(parts) >= 2:
            account_name = parts[0].strip()
            # Account number is typically the last part
            account_number = parts[-1].strip()
            return account_number, account_name

        return account_str, ""


class BankParserFactory:
    """Factory for creating appropriate bank parsers based on file type."""

    _parsers: list[type[BankParser]] = [
        CitiParser,
        UBSParser,
        MorganStanleyParser,
    ]

    @classmethod
    def get_parser(cls, file_path: str) -> BankParser | None:
        """Get the appropriate parser for a file.

        Args:
            file_path: Path to the file to parse.

        Returns:
            Appropriate BankParser instance or None if no parser matches.
        """
        path = Path(file_path)

        # First, check by extension for Excel files
        if path.suffix.lower() in (".xlsx", ".xls"):
            ms_parser = MorganStanleyParser()
            if ms_parser.can_parse(file_path):
                return ms_parser
            return None

        # For CSV files, try each parser
        if path.suffix.lower() == ".csv":
            for parser_cls in cls._parsers:
                csv_parser = parser_cls()
                if csv_parser.can_parse(file_path):
                    return csv_parser

        return None

    @classmethod
    def parse(cls, file_path: str) -> list[ParsedTransaction]:
        """Parse a file using the appropriate parser.

        Args:
            file_path: Path to the file to parse.

        Returns:
            List of ParsedTransaction objects.

        Raises:
            ValueError: If no parser can handle the file.
            FileNotFoundError: If the file does not exist.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        parser = cls.get_parser(file_path)
        if parser is None:
            raise ValueError(f"No parser available for file: {file_path}")

        return parser.parse(file_path)

    @classmethod
    def register_parser(cls, parser_cls: type[BankParser]) -> None:
        """Register a new parser class.

        Args:
            parser_cls: Parser class to register.
        """
        if parser_cls not in cls._parsers:
            cls._parsers.append(parser_cls)
